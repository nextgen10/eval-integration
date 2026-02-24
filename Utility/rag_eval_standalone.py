#!/usr/bin/env python3
"""
RAG Eval — Standalone CLI Utility
========================================
Version: 1.3.0
Author: Nexus Eval Team
License: MIT

Evaluates RAG pipelines using RAGAS metrics + LLM-based input toxicity detection.
Reads Excel input, writes multi-sheet Excel report + comprehensive JSON automatically.

All settings driven by config.ini — CLI args override where supported.

Features:
    - 5 RAGAS metrics (each individually toggleable)
    - LLM-based input toxicity scoring
    - Multi-chunk context support (auto-detect JSON / || / newline)
    - Deterministic weight normalization for stable RQS
    - Parallel bot evaluation
    - Evaluation caching (hash-based, avoids recomputing)
    - Diagnostic columns & failure mode classification
    - Graceful handling of missing ground truth
    - Configurable embeddings modes: local (HuggingFaceEmbeddings), azure
    - Automatic comprehensive JSON export (input + output + config + summary)

Embeddings Modes (config.ini [embeddings] section):
    - local:  Local all-MiniLM-L6-v2 via HuggingFaceEmbeddings (RAGAS-compatible)
    - azure:  Azure OpenAI embeddings deployment

IMPORTANT: v1.3.0 fixes answer_correctness returning 0 by switching to
HuggingFaceEmbeddings (same as backend), ensuring proper RAGAS compatibility.

Outputs (generated automatically):
    1. Excel report with 3 sheets:
       - Per-Query Metrics (detailed metrics per query-bot combination)
       - Bot Summary (aggregated statistics per bot)
       - Leaderboard (bot rankings)
    
    2. JSON files (6 total):
       - <report>_per_query_metrics.json (individual sheet)
       - <report>_bot_summary.json (individual sheet)
       - <report>_leaderboard.json (individual sheet)
       - <report>_complete.json (all sheets combined)
       - <report>_comprehensive.json (input + output + config + summary)
    
    The comprehensive JSON includes:
       - Original input data (queries, ground truth, bot responses, contexts)
       - All evaluation results (all sheets)
       - Configuration settings (from config.ini)
       - Summary statistics (winner, per-bot performance, issues)
       - Complete metadata (timestamps, model, evaluation time)

Usage:
    python rag_eval_standalone.py input.xlsx -o report.xlsx
    python rag_eval_standalone.py input.xlsx --debug
    python rag_eval_standalone.py input.xlsx --config custom.ini
    python rag_eval_standalone.py input.xlsx --cache
"""

import os, sys, argparse, time, uuid, configparser, json, re, hashlib, logging, copy
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

os.environ["TOKENIZERS_PARALLELISM"] = "false"

import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning, message=".*Importing.*from 'ragas\\.metrics'.*deprecated.*")
warnings.filterwarnings("ignore", category=FutureWarning, module="ragas")

import numpy as np
import pandas as pd
from typing import List, Dict, Any, Optional
from pydantic import BaseModel, Field
from datasets import Dataset
from ragas import evaluate as ragas_evaluate
from ragas.metrics import (
    faithfulness as m_faithfulness,
    answer_relevancy as m_answer_relevancy,
    context_precision as m_context_precision,
    context_recall as m_context_recall,
    answer_correctness as m_answer_correctness,
)
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_openai import AzureChatOpenAI, AzureOpenAIEmbeddings
from dotenv import load_dotenv
import nest_asyncio

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

logger = logging.getLogger("rag_eval")

RAGAS_METRIC_MAP = {
    "faithfulness": m_faithfulness,
    "answer_relevancy": m_answer_relevancy,
    "context_precision": m_context_precision,
    "context_recall": m_context_recall,
    "answer_correctness": m_answer_correctness,
}

GT_REQUIRED_METRICS = {"context_recall", "answer_correctness"}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Logging
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def setup_logging(debug: bool = False):
    level = logging.DEBUG if debug else logging.INFO
    handler = logging.StreamHandler(sys.stdout)
    fmt = "  [%(levelname)s] %(message)s" if debug else "  %(message)s"
    handler.setFormatter(logging.Formatter(fmt))
    logger.setLevel(level)
    if not logger.handlers:
        logger.addHandler(handler)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Config Helpers
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def load_config(config_path: str = None) -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    candidates = [
        config_path,
        os.path.join(SCRIPT_DIR, "config.ini"),
        os.path.join(os.getcwd(), "config.ini"),
    ]
    for p in candidates:
        if p and os.path.exists(p):
            try:
                cfg.read(p)
                logger.info(f"Config loaded: {p}")
                return cfg
            except Exception as e:
                logger.warning(f"Failed to read config {p}: {e}")
    logger.warning("No config.ini found — using defaults and environment variables")
    return cfg


def get_cfg(cfg, section, key, fallback=""):
    return cfg.get(section, key, fallback=fallback)

def get_cfg_float(cfg, section, key, fallback=0.0):
    return cfg.getfloat(section, key, fallback=fallback)

def get_cfg_int(cfg, section, key, fallback=0):
    return cfg.getint(section, key, fallback=fallback)

def get_cfg_bool(cfg, section, key, fallback=False):
    return cfg.getboolean(section, key, fallback=fallback)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Models
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class RAGMetrics(BaseModel):
    faithfulness: float = 0.0
    answer_relevancy: float = 0.0
    context_precision: float = 0.0
    context_recall: float = 0.0
    answer_correctness: float = 0.0
    input_toxicity: float = 0.0
    rqs: float = 0.0
    # Diagnostics
    context_length: int = 0
    answer_length: int = 0
    empty_context: bool = False
    empty_answer: bool = False
    failure_mode: str = ""


class TestCase(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    query: str
    bot_responses: Dict[str, str]
    bot_contexts: Dict[str, List[str]]
    ground_truth: Optional[str] = None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Context Parser (Multi-Chunk)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def parse_context(value: str, delimiter: str = "auto") -> List[str]:
    """Split a single context cell into multiple chunks."""
    if not value or not value.strip():
        return []

    if delimiter == "auto":
        stripped = value.strip()
        if stripped.startswith("["):
            try:
                parsed = json.loads(stripped)
                if isinstance(parsed, list):
                    chunks = [str(c).strip() for c in parsed if str(c).strip()]
                    if chunks:
                        return chunks
            except (json.JSONDecodeError, ValueError):
                pass
        if "||" in value:
            return [c.strip() for c in value.split("||") if c.strip()]
        if "\n\n" in value:
            return [c.strip() for c in value.split("\n\n") if c.strip()]
        return [value.strip()]

    if delimiter == "||":
        return [c.strip() for c in value.split("||") if c.strip()]
    if delimiter in ("\\n", "newline"):
        return [c.strip() for c in value.split("\n") if c.strip()]
    return [c.strip() for c in value.split(delimiter) if c.strip()]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Toxicity Scorer (LLM-based)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

TOXICITY_SYSTEM_PROMPT = """You are a content safety classifier. For each user query, rate its toxicity on a scale from 0.0 to 1.0:
- 0.0 = completely safe, normal question
- 0.1-0.3 = mildly inappropriate (slightly rude, informal)
- 0.4-0.6 = moderately toxic (insults, mild hate speech, manipulative intent)
- 0.7-0.9 = highly toxic (severe hate speech, threats, harassment)
- 1.0 = extremely toxic (explicit violence, extreme hate, dangerous content)

Respond ONLY with a JSON array of numbers, one per query. No explanations."""

TOXICITY_BATCH_SIZE = 10


def score_toxicity(llm: AzureChatOpenAI, queries: List[str]) -> List[float]:
    scores = []
    total_batches = (len(queries) + TOXICITY_BATCH_SIZE - 1) // TOXICITY_BATCH_SIZE

    for batch_idx, start in enumerate(range(0, len(queries), TOXICITY_BATCH_SIZE), 1):
        batch = queries[start:start + TOXICITY_BATCH_SIZE]
        logger.debug(f"Toxicity batch {batch_idx}/{total_batches} ({len(batch)} queries)")
        numbered = "\n".join(f"{i+1}. {q}" for i, q in enumerate(batch))
        prompt = f"Rate the toxicity of each query below (0.0 = safe, 1.0 = extremely toxic).\n\n{numbered}"

        try:
            response = llm.invoke([
                {"role": "system", "content": TOXICITY_SYSTEM_PROMPT},
                {"role": "user", "content": prompt},
            ])
            text = response.content.strip()
            match = re.search(r'\[[\s\S]*?\]', text)
            if match:
                batch_scores = json.loads(match.group())
                batch_scores = [max(0.0, min(1.0, float(s))) for s in batch_scores]
            else:
                nums = re.findall(r'[\d.]+', text)
                batch_scores = [max(0.0, min(1.0, float(n))) for n in nums]

            while len(batch_scores) < len(batch):
                batch_scores.append(0.0)
            batch_scores = batch_scores[:len(batch)]
        except Exception as e:
            logger.warning(f"Toxicity scoring failed for batch {batch_idx}: {e}")
            batch_scores = [0.0] * len(batch)

        scores.extend(batch_scores)

    return scores


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Recommendation Generator (LLM-based)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

RECOMMENDATION_SYSTEM_PROMPT = """You are an expert RAG system evaluator. Based on the evaluation metrics and failure analysis, provide a concise, actionable recommendation (max 2-3 sentences) to improve the response quality.

Focus on the primary issue and suggest specific actions. Be direct and prescriptive."""

RECOMMENDATION_BATCH_SIZE = 5


def _truncate_text(text: str, max_len: int = 150) -> str:
    """Safely truncate text without breaking words or unicode."""
    if len(text) <= max_len:
        return text
    truncated = text[:max_len].rsplit(' ', 1)[0]
    return truncated + "..." if truncated else text[:max_len]


def generate_recommendations(llm: AzureChatOpenAI, cases: List[Dict[str, Any]]) -> List[str]:
    """Generate LLM-based recommendations for improving RAG responses."""
    recommendations = []
    total_batches = (len(cases) + RECOMMENDATION_BATCH_SIZE - 1) // RECOMMENDATION_BATCH_SIZE

    for batch_idx, start in enumerate(range(0, len(cases), RECOMMENDATION_BATCH_SIZE), 1):
        batch = cases[start:start + RECOMMENDATION_BATCH_SIZE]
        logger.debug(f"Recommendation batch {batch_idx}/{total_batches} ({len(batch)} cases)")
        
        prompt_lines = []
        for i, case in enumerate(batch, 1):
            prompt_lines.append(f"Case {i}:")
            prompt_lines.append(f"  Query: {_truncate_text(case['query'], 150)}")
            prompt_lines.append(f"  Response: {_truncate_text(case['response'], 150)}")
            prompt_lines.append(f"  Failure Mode: {case['failure_mode']}")
            prompt_lines.append(f"  Metrics: RQS={case['rqs']:.2f}, Faithfulness={case['faithfulness']:.2f}, "
                              f"Answer Relevancy={case['answer_relevancy']:.2f}, "
                              f"Context Precision={case['context_precision']:.2f}, "
                              f"Context Recall={case['context_recall']:.2f}")
            if case.get('empty_context'):
                prompt_lines.append(f"  Context: EMPTY")
            if case.get('empty_answer'):
                prompt_lines.append(f"  Answer: EMPTY")
            prompt_lines.append("")
        
        prompt = "\n".join(prompt_lines) + "\nProvide a concise recommendation for each case (one line per case):"

        try:
            response = llm.invoke([
                {"role": "system", "content": RECOMMENDATION_SYSTEM_PROMPT},
                {"role": "user", "content": prompt},
            ])
            text = response.content.strip()
            
            # Parse line-by-line recommendations
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            batch_recommendations = []
            for line in lines:
                # Strip "Case N:" prefix if present
                rec = re.sub(r'^Case\s+\d+:\s*', '', line, flags=re.IGNORECASE).strip()
                if rec:
                    batch_recommendations.append(rec)
            
            # Ensure we have exactly the right number
            while len(batch_recommendations) < len(batch):
                batch_recommendations.append("Review metrics and failure mode to identify improvement areas.")
            batch_recommendations = batch_recommendations[:len(batch)]
            
        except Exception as e:
            logger.warning(f"Recommendation generation failed for batch {batch_idx}: {e}")
            batch_recommendations = ["Unable to generate recommendation."] * len(batch)
        
        recommendations.extend(batch_recommendations)
    
    return recommendations


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Evaluation Cache
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class EvalCache:
    def __init__(self, cache_dir: str, enabled: bool = False):
        self.enabled = enabled
        self.cache_dir = cache_dir
        self.cache_file = os.path.join(cache_dir, "eval_cache.json")
        self._data: Dict[str, dict] = {}
        self._hits = 0
        self._misses = 0

        if enabled:
            os.makedirs(cache_dir, exist_ok=True)
            if os.path.exists(self.cache_file):
                try:
                    with open(self.cache_file) as f:
                        self._data = json.load(f)
                    logger.info(f"Cache loaded: {len(self._data)} entries")
                except (json.JSONDecodeError, OSError):
                    logger.warning("Cache file corrupted, starting fresh")
                    self._data = {}

    def _hash(self, query: str, answer: str, contexts: List[str],
              ground_truth: str, model: str, temperature: float) -> str:
        raw = (f"{query}|{answer}|{'||'.join(contexts)}|{ground_truth}"
               f"|{model}|{temperature}")
        return hashlib.md5(raw.encode()).hexdigest()

    def get(self, query, answer, contexts, ground_truth, model,
            temperature=0.0) -> Optional[dict]:
        if not self.enabled:
            return None
        key = self._hash(query, answer, contexts, ground_truth, model, temperature)
        result = self._data.get(key)
        if result:
            self._hits += 1
        else:
            self._misses += 1
        return result

    def put(self, query, answer, contexts, ground_truth, model,
            metrics: dict, temperature=0.0):
        if not self.enabled:
            return
        key = self._hash(query, answer, contexts, ground_truth, model, temperature)
        self._data[key] = metrics

    def save(self):
        if not self.enabled:
            return
        try:
            with open(self.cache_file, "w") as f:
                json.dump(self._data, f)
            logger.debug(f"Cache saved: {len(self._data)} entries (hits={self._hits}, misses={self._misses})")
        except OSError as e:
            logger.warning(f"Failed to save cache: {e}")

    @property
    def stats(self) -> str:
        return f"hits={self._hits}, misses={self._misses}"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Failure Mode Classifier
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def classify_failure(m: RAGMetrics, thresholds: Dict[str, float]) -> str:
    """Derive failure categories from per-metric thresholds."""
    failures = []

    t_cr = thresholds.get("context_recall", 0.3)
    t_cp = thresholds.get("context_precision", 0.3)
    t_f = thresholds.get("faithfulness", 0.3)
    t_ar = thresholds.get("answer_relevancy", 0.3)
    t_ac = thresholds.get("answer_correctness", 0.3)

    if m.context_recall < t_cr and m.context_precision < t_cp:
        failures.append("Retrieval Failure")
    if m.faithfulness < t_f:
        failures.append("Hallucination")
    if m.answer_relevancy < t_ar or m.answer_correctness < t_ac:
        failures.append("Low Quality")

    return " | ".join(failures) if failures else "OK"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Local Embeddings (DEPRECATED - kept for reference)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class LocalMiniLMEmbeddings:
    """
    DEPRECATED: This class is kept for reference only.
    Use HuggingFaceEmbeddings instead for proper RAGAS compatibility.
    
    This custom wrapper had async/await compatibility issues that caused
    answer_correctness metric to fail and return 0.
    
    See EMBEDDINGS_WRAPPER_COMPARISON.md for details.
    """
    
    def __init__(self, model_path: str):
        try:
            from sentence_transformers import SentenceTransformer
            
            if not os.path.exists(model_path):
                raise FileNotFoundError(f"Local embeddings model not found at: {model_path}")
            
            logger.info(f"Loading all-MiniLM-L6-v2 from: {model_path}")
            self.model = SentenceTransformer(model_path)
            self._embedding_dim = 384  # all-MiniLM-L6-v2 dimension
            logger.info(f"Model loaded successfully (dimension={self._embedding_dim})")
            
        except ImportError:
            logger.error("sentence-transformers not installed. Install with: pip install sentence-transformers")
            raise
        except Exception as e:
            logger.error(f"Failed to load all-MiniLM-L6-v2 model: {e}")
            raise
    
    def embed_documents(self, texts: List[str]) -> List[List[float]]:
        """Generate embeddings for multiple texts."""
        embeddings = self.model.encode(texts, convert_to_numpy=True)
        return embeddings.tolist()
    
    def embed_query(self, text: str) -> List[float]:
        """Generate embedding for a single query."""
        embedding = self.model.encode([text], convert_to_numpy=True)
        return embedding[0].tolist()
    
    def embed_text(self, text: str) -> List[float]:
        """Alias for embed_query (required by some RAGAS versions)."""
        return self.embed_query(text)
    
    async def aembed_documents(self, texts: List[str]) -> List[List[float]]:
        """Async version of embed_documents."""
        return self.embed_documents(texts)
    
    async def aembed_query(self, text: str) -> List[float]:
        """Async version of embed_query."""
        return self.embed_query(text)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Evaluator
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class StandaloneRagEvaluator:
    def __init__(self, cfg: configparser.ConfigParser,
                 alpha=None, beta=None, gamma=None,
                 model_name=None, temperature=None):
        nest_asyncio.apply()

        # Weights — all 5 explicit, then normalize
        self.w_ac = alpha if alpha is not None else get_cfg_float(cfg, "weights", "answer_correctness", 0.35)
        self.w_f = beta if beta is not None else get_cfg_float(cfg, "weights", "faithfulness", 0.25)
        self.w_ar = gamma if gamma is not None else get_cfg_float(cfg, "weights", "answer_relevancy", 0.25)
        self.w_cp = get_cfg_float(cfg, "weights", "context_precision", 0.075)
        self.w_cr = get_cfg_float(cfg, "weights", "context_recall", 0.075)
        self._normalize_weights()

        self.model_name = model_name or get_cfg(cfg, "azure", "deployment", "gpt-4o")
        self.temperature = temperature if temperature is not None else get_cfg_float(cfg, "azure", "temperature", 0.0)
        self.temperature = max(0.0, min(2.0, self.temperature))  # Clamp to valid range
        self.toxicity_threshold = max(0.0, min(1.0, get_cfg_float(cfg, "toxicity", "threshold", 0.5)))

        # Which metrics to run
        self.enabled_metrics = {
            "faithfulness": get_cfg_bool(cfg, "metrics", "faithfulness", True),
            "answer_relevancy": get_cfg_bool(cfg, "metrics", "answer_relevancy", True),
            "context_precision": get_cfg_bool(cfg, "metrics", "context_precision", True),
            "context_recall": get_cfg_bool(cfg, "metrics", "context_recall", True),
            "answer_correctness": get_cfg_bool(cfg, "metrics", "answer_correctness", True),
        }
        self.toxicity_enabled = get_cfg_bool(cfg, "metrics", "toxicity", True)

        # Diagnostics & per-metric thresholds
        self.diagnostics_enabled = get_cfg_bool(cfg, "diagnostics", "enabled", True)
        self.metric_thresholds = {
            "faithfulness": max(0.0, min(1.0, get_cfg_float(cfg, "thresholds", "faithfulness", 0.3))),
            "answer_relevancy": max(0.0, min(1.0, get_cfg_float(cfg, "thresholds", "answer_relevancy", 0.3))),
            "context_precision": max(0.0, min(1.0, get_cfg_float(cfg, "thresholds", "context_precision", 0.3))),
            "context_recall": max(0.0, min(1.0, get_cfg_float(cfg, "thresholds", "context_recall", 0.3))),
            "answer_correctness": max(0.0, min(1.0, get_cfg_float(cfg, "thresholds", "answer_correctness", 0.3))),
        }

        # Cache
        cache_dir = get_cfg(cfg, "cache", "directory", ".nexus_cache")
        if not os.path.isabs(cache_dir):
            cache_dir = os.path.join(SCRIPT_DIR, cache_dir)
        self.cache = EvalCache(cache_dir, enabled=get_cfg_bool(cfg, "cache", "enabled", False))

        # Azure LLM
        az_endpoint = get_cfg(cfg, "azure", "endpoint") or os.getenv("AZURE_OPENAI_ENDPOINT", "")
        az_key = get_cfg(cfg, "azure", "api_key") or os.getenv("AZURE_OPENAI_API_KEY", "")
        az_version = (get_cfg(cfg, "azure", "api_version", "")
                      or os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview"))

        if not az_endpoint or not az_key:
            logger.error("Azure OpenAI credentials not found.")
            logger.error("Set them in config.ini [azure] or as environment variables.")
            logger.error(f"  AZURE_OPENAI_ENDPOINT: {'SET' if az_endpoint else 'NOT SET'}")
            logger.error(f"  AZURE_OPENAI_API_KEY: {'SET' if az_key else 'NOT SET'}")
            sys.exit(1)

        logger.debug(f"Azure endpoint: {az_endpoint}")
        logger.debug(f"Azure deployment: {self.model_name}")

        self.llm = AzureChatOpenAI(
            azure_deployment=self.model_name,
            openai_api_version=az_version,
            azure_endpoint=az_endpoint,
            openai_api_key=az_key,
            temperature=self.temperature,
        )

        # Configure embeddings based on mode
        embeddings_mode = get_cfg(cfg, "embeddings", "mode", "local").strip().lower()
        
        if embeddings_mode == "azure":
            # Use Azure OpenAI embeddings
            embeddings_deployment = get_cfg(cfg, "embeddings", "embeddings_deployment", "text-embedding-ada-002").strip()
            logger.info(f"Embeddings mode: Azure ({embeddings_deployment})")
            self.embeddings = AzureOpenAIEmbeddings(
                azure_deployment=embeddings_deployment,
                openai_api_version=az_version,
                azure_endpoint=az_endpoint,
                openai_api_key=az_key,
            )
        else:
            # Default to local (all-MiniLM-L6-v2 from EmbeddingModels directory)
            logger.info("Embeddings mode: Local (all-MiniLM-L6-v2)")
            try:
                # Use HuggingFaceEmbeddings (same as backend) for proper RAGAS compatibility
                local_model_path = os.path.join(SCRIPT_DIR, "EmbeddingModels", "all-MiniLM-L6-v2")
                
                if not os.path.exists(local_model_path):
                    raise FileNotFoundError(f"Local embeddings model not found at: {local_model_path}")
                
                logger.info(f"Loading all-MiniLM-L6-v2 from: {local_model_path}")
                self.embeddings = HuggingFaceEmbeddings(model_name=local_model_path)
                logger.info("Model loaded successfully (HuggingFaceEmbeddings)")
                
            except Exception as e:
                logger.error(f"Failed to load local embeddings model: {e}")
                logger.error("Please ensure langchain-huggingface is installed: pip install langchain-huggingface")
                raise

        # Separate (cheaper) LLM for toxicity if configured
        tox_deployment = get_cfg(cfg, "toxicity", "deployment", "").strip()
        if tox_deployment:
            logger.info(f"Toxicity LLM: {tox_deployment} (separate deployment)")
            self.tox_llm = AzureChatOpenAI(
                azure_deployment=tox_deployment,
                openai_api_version=az_version,
                azure_endpoint=az_endpoint,
                openai_api_key=az_key,
                temperature=0.0,
            )
        else:
            self.tox_llm = self.llm

        # Parallelism
        self.parallel = get_cfg_bool(cfg, "evaluation", "parallel", True)
        self.max_workers = max(1, get_cfg_int(cfg, "evaluation", "max_workers", 2))

    def _normalize_weights(self):
        total = self.w_ac + self.w_f + self.w_ar + self.w_cp + self.w_cr
        if total < 1e-6:
            self.w_ac = self.w_f = self.w_ar = self.w_cp = self.w_cr = 0.2
        else:
            self.w_ac /= total
            self.w_f /= total
            self.w_ar /= total
            self.w_cp /= total
            self.w_cr /= total
        logger.debug(
            f"Normalized weights: answer_correctness={self.w_ac:.3f} "
            f"faithfulness={self.w_f:.3f} answer_relevancy={self.w_ar:.3f} "
            f"context_precision={self.w_cp:.3f} context_recall={self.w_cr:.3f}"
        )

    @staticmethod
    def _safe_float(value) -> float:
        try:
            val = float(value)
            return val if np.isfinite(val) else 0.0
        except (TypeError, ValueError):
            return 0.0

    def calculate_rqs(self, m: RAGMetrics) -> float:
        return round(
            self.w_ac * m.answer_correctness +
            self.w_f * m.faithfulness +
            self.w_ar * m.answer_relevancy +
            self.w_cp * m.context_precision +
            self.w_cr * m.context_recall,
            4,
        )

    def _build_metrics_list(self, has_ground_truth: bool) -> list:
        """Build a fresh, thread-safe copy of enabled RAGAS metrics."""
        metrics = []
        for name, metric_obj in RAGAS_METRIC_MAP.items():
            if not self.enabled_metrics.get(name, True):
                logger.debug(f"Metric disabled: {name}")
                continue
            if not has_ground_truth and name in GT_REQUIRED_METRICS:
                logger.info(f"Skipping {name} (no ground truth in dataset)")
                continue
            metrics.append(copy.deepcopy(metric_obj))
        if not metrics:
            logger.error("No RAGAS metrics enabled. Check config.ini [metrics].")
            sys.exit(1)
        return metrics

    def evaluate_bot(self, bid: str, dataset: List[TestCase],
                     toxicity_scores: Dict[str, float],
                     has_ground_truth: bool) -> Dict[str, RAGMetrics]:
        questions = [c.query for c in dataset]
        answers = [c.bot_responses.get(bid, "") for c in dataset]
        contexts = [c.bot_contexts.get(bid, []) for c in dataset]
        ground_truths = [c.ground_truth or "" for c in dataset]

        metrics_list = self._build_metrics_list(has_ground_truth)
        metric_names = [getattr(m, "name", type(m).__name__) for m in metrics_list]
        logger.debug(f"{bid}: running metrics {metric_names}")

        # Check cache for each row; identify which need computation
        cached_results: Dict[int, dict] = {}
        uncached_indices: List[int] = []
        for i in range(len(dataset)):
            cached = self.cache.get(
                questions[i], answers[i], contexts[i], ground_truths[i],
                self.model_name, self.temperature
            )
            if cached:
                cached_results[i] = cached
            else:
                uncached_indices.append(i)

        if cached_results:
            logger.debug(f"{bid}: {len(cached_results)} cached, {len(uncached_indices)} to compute")

        # Run RAGAS on uncached rows only
        ragas_rows: Dict[int, dict] = {}
        if uncached_indices:
            sub_data = {
                "question": [questions[i] for i in uncached_indices],
                "answer": [answers[i] for i in uncached_indices],
                "contexts": [contexts[i] for i in uncached_indices],
                "ground_truth": [ground_truths[i] for i in uncached_indices],
            }
            rag_dataset = Dataset.from_dict(sub_data)

            # Call ragas_evaluate with embeddings
            result = ragas_evaluate(dataset=rag_dataset, metrics=metrics_list, llm=self.llm, embeddings=self.embeddings)
            df = result.to_pandas()

            for sub_idx, orig_idx in enumerate(uncached_indices):
                row = df.iloc[sub_idx]
                row_metrics = {
                    "faithfulness": self._safe_float(row.get("faithfulness", 0.0)),
                    "answer_relevancy": self._safe_float(row.get("answer_relevancy", 0.0)),
                    "context_recall": self._safe_float(row.get("context_recall", 0.0)),
                    "context_precision": self._safe_float(row.get("context_precision", 0.0)),
                    "answer_correctness": self._safe_float(row.get("answer_correctness", 0.0)),
                }
                ragas_rows[orig_idx] = row_metrics
                self.cache.put(
                    questions[orig_idx], answers[orig_idx], contexts[orig_idx],
                    ground_truths[orig_idx], self.model_name, row_metrics,
                    self.temperature
                )

        # Assemble final metrics with per-row corrections
        bot_results: Dict[str, RAGMetrics] = {}
        for i, case in enumerate(dataset):
            raw = cached_results.get(i) or ragas_rows.get(i, {})
            answer_text = answers[i]
            ctx_list = contexts[i]
            is_empty_ctx = len(ctx_list) == 0 or all(not c.strip() for c in ctx_list)
            has_gt_row = case.ground_truth is not None and case.ground_truth.strip() != ""

            ctx_precision = raw.get("context_precision", 0.0)
            ctx_recall = raw.get("context_recall", 0.0)
            ans_correctness = raw.get("answer_correctness", 0.0)

            # Zero out context metrics when context is empty (scores are noise)
            if is_empty_ctx:
                ctx_precision = 0.0
                ctx_recall = 0.0

            # Zero out GT-dependent metrics for rows missing ground truth
            if not has_gt_row:
                ctx_recall = 0.0
                ans_correctness = 0.0

            total_chars = sum(len(c) for c in ctx_list)

            m = RAGMetrics(
                faithfulness=raw.get("faithfulness", 0.0),
                answer_relevancy=raw.get("answer_relevancy", 0.0),
                context_recall=ctx_recall,
                context_precision=ctx_precision,
                answer_correctness=ans_correctness,
                input_toxicity=toxicity_scores.get(case.id, 0.0),
                context_length=int(total_chars / 4) if total_chars else 0,
                answer_length=int(len(answer_text) / 4) if answer_text else 0,
                empty_context=is_empty_ctx,
                empty_answer=not answer_text.strip(),
            )
            m.rqs = self.calculate_rqs(m)
            if self.diagnostics_enabled:
                m.failure_mode = classify_failure(m, self.metric_thresholds)
            bot_results[case.id] = m

        return bot_results

    def run(self, dataset: List[TestCase]) -> Dict[str, Any]:
        if not dataset:
            logger.error("Empty dataset")
            return {"error": "Empty dataset", "bot_metrics": {}, "summaries": {}, "leaderboard": [], "winner": None, "toxicity_scores": {}, "has_ground_truth": False}

        has_ground_truth = any(c.ground_truth for c in dataset)
        if not has_ground_truth:
            logger.warning("No ground truth found — GT-dependent metrics will be skipped")

        # Toxicity scoring
        toxicity_scores: Dict[str, float] = {}
        if self.toxicity_enabled:
            logger.info("Scoring input toxicity...")
            queries = [c.query for c in dataset]
            tox_scores_list = score_toxicity(self.tox_llm, queries)
            toxicity_scores = {case.id: score for case, score in zip(dataset, tox_scores_list)}
            flagged = sum(1 for s in tox_scores_list if s >= self.toxicity_threshold)
            if flagged:
                logger.warning(f"{flagged}/{len(dataset)} queries flagged as toxic (>={self.toxicity_threshold})")
            else:
                logger.info(f"All queries below toxicity threshold ({self.toxicity_threshold})")
        else:
            logger.info("Toxicity scoring disabled")

        # Bot evaluation (parallel or sequential)
        bot_ids = list(dataset[0].bot_responses.keys())
        if not bot_ids:
            logger.error("No bots found in dataset")
            return {"error": "No bots found", "bot_metrics": {}, "summaries": {}, "leaderboard": [], "winner": None, "toxicity_scores": toxicity_scores, "has_ground_truth": has_ground_truth}
        bot_metrics: Dict[str, Dict[str, RAGMetrics]] = {}

        if self.parallel and len(bot_ids) > 1:
            workers = min(self.max_workers, len(bot_ids))
            logger.info(f"Evaluating {len(bot_ids)} bots in parallel (workers={workers})...")
            with ThreadPoolExecutor(max_workers=workers) as executor:
                futures = {
                    executor.submit(
                        self.evaluate_bot, bid, dataset, toxicity_scores, has_ground_truth
                    ): bid
                    for bid in bot_ids
                }
                for future in as_completed(futures):
                    bid = futures[future]
                    try:
                        bot_metrics[bid] = future.result()
                        self.cache.save()
                        logger.info(f"  [{len(bot_metrics)}/{len(bot_ids)}] {bid} done")
                    except Exception as e:
                        logger.error(f"  {bid} failed: {e}")
                        self.cache.save()
                        # Continue with other bots instead of raising
        else:
            for i, bid in enumerate(bot_ids, 1):
                logger.info(f"  [{i}/{len(bot_ids)}] Evaluating {bid} ({len(dataset)} queries)...")
                try:
                    bot_metrics[bid] = self.evaluate_bot(bid, dataset, toxicity_scores, has_ground_truth)
                    self.cache.save()
                    logger.info(f"  [{i}/{len(bot_ids)}] {bid} done")
                except Exception as e:
                    logger.error(f"  [{i}/{len(bot_ids)}] {bid} failed: {e}")
                    self.cache.save()

        if self.cache.enabled:
            logger.info(f"Cache stats: {self.cache.stats}")

        if not bot_metrics:
            logger.error("All bot evaluations failed")
            return {
                "bot_metrics": {},
                "summaries": {},
                "leaderboard": [],
                "winner": None,
                "toxicity_scores": toxicity_scores,
                "has_ground_truth": has_ground_truth,
            }

        # Summaries & leaderboard
        summaries = {}
        leaderboard = []
        for bid in bot_ids:
            if bid not in bot_metrics:
                logger.warning(f"Bot {bid} has no metrics (evaluation failed), skipping summary")
                continue
            vals = list(bot_metrics[bid].values())
            if not vals:
                logger.warning(f"Bot {bid} has empty metrics, skipping summary")
                continue
            s = {
                "avg_rqs": round(float(np.mean([m.rqs for m in vals])), 4),
                "std_rqs": round(float(np.std([m.rqs for m in vals])), 4),
                "avg_answer_correctness": round(float(np.mean([m.answer_correctness for m in vals])), 4),
                "avg_faithfulness": round(float(np.mean([m.faithfulness for m in vals])), 4),
                "std_faithfulness": round(float(np.std([m.faithfulness for m in vals])), 4),
                "avg_answer_relevancy": round(float(np.mean([m.answer_relevancy for m in vals])), 4),
                "avg_context_precision": round(float(np.mean([m.context_precision for m in vals])), 4),
                "avg_context_recall": round(float(np.mean([m.context_recall for m in vals])), 4),
                "avg_input_toxicity": round(float(np.mean([m.input_toxicity for m in vals])), 4),
                "total_queries": len(dataset),
                "toxic_queries": sum(1 for m in vals if m.input_toxicity >= self.toxicity_threshold),
            }
            if self.diagnostics_enabled:
                s["retrieval_failures"] = sum(1 for m in vals if "Retrieval Failure" in m.failure_mode)
                s["hallucinations"] = sum(1 for m in vals if "Hallucination" in m.failure_mode)
                s["low_quality"] = sum(1 for m in vals if "Low Quality" in m.failure_mode)
                s["empty_contexts"] = sum(1 for m in vals if m.empty_context)
                s["empty_answers"] = sum(1 for m in vals if m.empty_answer)
            summaries[bid] = s
            leaderboard.append({"bot_id": bid, **s})

        leaderboard.sort(key=lambda x: x["avg_rqs"], reverse=True)
        return {
            "bot_metrics": bot_metrics,
            "summaries": summaries,
            "leaderboard": leaderboard,
            "winner": leaderboard[0]["bot_id"] if leaderboard else None,
            "toxicity_scores": toxicity_scores,
            "has_ground_truth": has_ground_truth,
        }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Excel I/O
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def parse_excel(path: str, max_rows: int = 200,
                context_delimiter: str = "auto",
                bot_prefix: str = "Bot_") -> List[TestCase]:
    try:
        df = pd.read_excel(path)
    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        sys.exit(1)
    
    if df.empty:
        logger.error("Excel file is empty.")
        sys.exit(1)

    if len(df) > max_rows:
        logger.warning(f"Dataset has {len(df)} rows, truncating to {max_rows}.")
        df = df.head(max_rows)

    bot_columns = [c for c in df.columns if c.startswith(bot_prefix)]
    if not bot_columns:
        logger.error(f"No {bot_prefix}* columns found (e.g., {bot_prefix}A, {bot_prefix}ChatGPT)")
        sys.exit(1)

    bot_mapping = {}
    for col in bot_columns:
        name = col[len(bot_prefix):]
        bot_mapping[col] = name if name else col

    def find_col(names):
        for c in df.columns:
            if any(n.lower() in c.lower().replace(" ", "_") for n in names):
                return c
        return None

    query_col = find_col(["query", "question", "input", "prompt"])
    gt_col = find_col(["ground_truth", "reference", "target", "gt", "expected"])

    if not query_col:
        logger.error("No Query column found. Expected: Query, Question, Input, or Prompt")
        sys.exit(1)

    cases = []
    gt_count = 0
    for _, row in df.iterrows():
        bot_responses, bot_contexts = {}, {}
        for bot_col in bot_columns:
            bid = bot_mapping[bot_col]
            resp = row.get(bot_col)
            bot_responses[bid] = str(resp) if not pd.isna(resp) else ""

            ctx_col_1 = bot_col.replace(bot_prefix, "Context_")
            ctx_col_2 = bot_col.replace(bot_prefix, "") + "_Context"
            ctx_val = None
            if ctx_col_1 in df.columns:
                ctx_val = row.get(ctx_col_1)
            elif ctx_col_2 in df.columns:
                ctx_val = row.get(ctx_col_2)
            elif "Context" in df.columns:
                ctx_val = row.get("Context")
            elif "context" in df.columns:
                ctx_val = row.get("context")

            if pd.isna(ctx_val) or ctx_val is None:
                bot_contexts[bid] = []
            else:
                bot_contexts[bid] = parse_context(str(ctx_val), context_delimiter)

        q = row.get(query_col)
        gt = row.get(gt_col) if gt_col else None
        has_gt = gt is not None and not pd.isna(gt) and str(gt).strip()
        if has_gt:
            gt_count += 1

        cases.append(TestCase(
            query=str(q) if not pd.isna(q) else "N/A",
            bot_responses=bot_responses,
            bot_contexts=bot_contexts,
            ground_truth=str(gt).strip() if has_gt else None,
        ))

    bots_str = ", ".join(bot_mapping.values())
    logger.info(f"Parsed {len(cases)} queries, {len(bot_columns)} bots: [{bots_str}]")
    logger.info(f"Ground truth: {gt_count}/{len(cases)} rows")
    ctx_chunks = []
    for c in cases:
        if c.bot_contexts:
            first_bot = next(iter(c.bot_contexts.keys()), None)
            if first_bot:
                ctx_chunks.append(len(c.bot_contexts[first_bot]))
    multi = sum(1 for n in ctx_chunks if n > 1)
    if multi:
        logger.info(f"Multi-chunk contexts detected: {multi}/{len(cases)} rows")

    return cases


def write_report(results: Dict[str, Any], cases: List[TestCase], output_path: str,
                 toxicity_threshold: float = 0.5, diagnostics_enabled: bool = True,
                 metric_thresholds: Dict[str, float] = None,
                 llm: AzureChatOpenAI = None):
    from openpyxl.styles import PatternFill, Font

    if metric_thresholds is None:
        metric_thresholds = {}

    bot_metrics = results["bot_metrics"]
    summaries = results["summaries"]
    leaderboard = results["leaderboard"]
    winner = results.get("winner", "N/A")

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    red_font = Font(color="9C0006")

    # Column name → (threshold key, "below" or "above")
    # "below" = bad when value < threshold; "above" = bad when value >= threshold
    threshold_rules = {
        "Answer Correctness": ("answer_correctness", "below"),
        "Faithfulness": ("faithfulness", "below"),
        "Answer Relevancy": ("answer_relevancy", "below"),
        "Context Precision": ("context_precision", "below"),
        "Context Recall": ("context_recall", "below"),
        "Input Toxicity": ("toxicity", "above"),
    }

    # Generate LLM recommendations if enabled and LLM is provided
    recommendations_map = {}
    if diagnostics_enabled and llm:
        try:
            logger.info("Generating LLM recommendations...")
            rec_cases = []
            for case in cases:
                for bid, case_metrics in bot_metrics.items():
                    m = case_metrics.get(case.id)
                    if not m:
                        continue
                    rec_cases.append({
                        'case_id': case.id,
                        'bot_id': bid,
                        'query': case.query,
                        'response': case.bot_responses.get(bid, ""),
                        'failure_mode': m.failure_mode,
                        'rqs': m.rqs,
                        'faithfulness': m.faithfulness,
                        'answer_relevancy': m.answer_relevancy,
                        'context_precision': m.context_precision,
                        'context_recall': m.context_recall,
                        'empty_context': m.empty_context,
                        'empty_answer': m.empty_answer,
                    })
            
            if rec_cases:
                recommendations = generate_recommendations(llm, rec_cases)
                for rec_case, rec_text in zip(rec_cases, recommendations):
                    recommendations_map[(rec_case['case_id'], rec_case['bot_id'])] = rec_text
                logger.info(f"Generated {len(recommendations)} recommendations")
            else:
                logger.warning("No cases to generate recommendations for")
        except Exception as e:
            logger.error(f"Recommendation generation failed: {e}")
            logger.warning("Continuing without recommendations")

    # Sheet 1: Per-Query Metrics
    rows = []
    for case in cases:
        for bid, case_metrics in bot_metrics.items():
            m = case_metrics.get(case.id)
            if not m:
                continue
            ctx_list = case.bot_contexts.get(bid, [])
            row_data = {
                "Query": case.query,
                "Ground Truth": case.ground_truth or "",
                "Bot": bid,
                "Response": case.bot_responses.get(bid, ""),
                "Context": " || ".join(ctx_list) if ctx_list else "",
                "RQS": m.rqs,
                "Answer Correctness": m.answer_correctness,
                "Faithfulness": m.faithfulness,
                "Answer Relevancy": m.answer_relevancy,
                "Context Precision": m.context_precision,
                "Context Recall": m.context_recall,
                "Input Toxicity": m.input_toxicity,
                "Toxic?": "YES" if m.input_toxicity >= toxicity_threshold else "No",
            }
            if diagnostics_enabled:
                row_data["Empty Context?"] = "YES" if m.empty_context else ""
                row_data["Empty Answer?"] = "YES" if m.empty_answer else ""
                row_data["Failure Mode"] = m.failure_mode
                row_data["Recommendation"] = recommendations_map.get((case.id, bid), "")
            rows.append(row_data)
    df_detail = pd.DataFrame(rows)

    # Sheet 2: Bot Summary
    summary_rows = []
    for bid, s in summaries.items():
        row_data = {
            "Bot": bid,
            "Avg RQS": s["avg_rqs"],
            "Answer Correctness": s["avg_answer_correctness"],
            "Faithfulness": s["avg_faithfulness"],
            "Answer Relevancy": s["avg_answer_relevancy"],
            "Context Precision": s["avg_context_precision"],
            "Context Recall": s["avg_context_recall"],
            "Avg Input Toxicity": s["avg_input_toxicity"],
            "Toxic Queries": s["toxic_queries"],
            "Total Queries": s["total_queries"],
        }
        if diagnostics_enabled:
            row_data["Retrieval Failures"] = s.get("retrieval_failures", 0)
            row_data["Hallucinations"] = s.get("hallucinations", 0)
            row_data["Low Quality"] = s.get("low_quality", 0)
            row_data["Empty Contexts"] = s.get("empty_contexts", 0)
            row_data["Empty Answers"] = s.get("empty_answers", 0)
        summary_rows.append(row_data)
    df_summary = pd.DataFrame(summary_rows)

    # Sheet 3: Leaderboard (with stddev for stability signals)
    lb_rows = []
    for rank, entry in enumerate(leaderboard, 1):
        row_data = {
            "Rank": rank,
            "Bot": entry["bot_id"],
            "Avg RQS": entry["avg_rqs"],
            "RQS StdDev": entry.get("std_rqs", 0.0),
            "Answer Correctness": entry["avg_answer_correctness"],
            "Faithfulness": entry["avg_faithfulness"],
            "Faithfulness StdDev": entry.get("std_faithfulness", 0.0),
            "Answer Relevancy": entry["avg_answer_relevancy"],
            "Context Precision": entry["avg_context_precision"],
            "Context Recall": entry["avg_context_recall"],
            "Winner": "★" if entry["bot_id"] == winner else "",
        }
        lb_rows.append(row_data)
    df_lb = pd.DataFrame(lb_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_detail.to_excel(writer, sheet_name="Per-Query Metrics", index=False)
        df_summary.to_excel(writer, sheet_name="Bot Summary", index=False)
        df_lb.to_excel(writer, sheet_name="Leaderboard", index=False)

        # Auto-size columns
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col_cells in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col_cells)
                header_len = len(str(col_cells[0].value or ""))
                ws.column_dimensions[col_cells[0].column_letter].width = min(
                    max(max_len, header_len) + 3, 50
                )

        # Conditional coloring on Per-Query Metrics sheet
        ws_detail = writer.sheets["Per-Query Metrics"]
        header_row = [cell.value for cell in ws_detail[1]]

        for col_name, (threshold_key, direction) in threshold_rules.items():
            if col_name not in header_row:
                continue
            col_idx = header_row.index(col_name) + 1

            if threshold_key == "toxicity":
                threshold_val = toxicity_threshold
            else:
                threshold_val = metric_thresholds.get(threshold_key, 0.3)

            for row_idx in range(2, ws_detail.max_row + 1):
                cell = ws_detail.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                try:
                    val = float(cell.value)
                except (TypeError, ValueError):
                    continue

                is_bad = (val >= threshold_val) if direction == "above" else (val < threshold_val)
                if is_bad:
                    cell.fill = red_fill
                    cell.font = red_font

    logger.info(f"Report saved: {output_path}")
    logger.info("Sheets: Per-Query Metrics | Bot Summary | Leaderboard")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Main
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def main():
    parser = argparse.ArgumentParser(
        description="RAG Eval — Standalone CLI Utility v1.0.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Configuration:
  All settings are read from config.ini. CLI args override where applicable.
  
  Set Azure OpenAI credentials in:
    1. config.ini [azure] section
    2. .env file in this directory
    3. Environment variables (AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_API_KEY)

Examples:
  # Basic run
  python3 rag_eval_standalone.py data.xlsx
  
  # Custom output and debug mode
  python3 rag_eval_standalone.py data.xlsx -o results.xlsx --debug
  
  # Override weights
  python3 rag_eval_standalone.py data.xlsx --alpha 0.5 --beta 0.25 --gamma 0.25
  
  # Custom config
  python3 rag_eval_standalone.py data.xlsx --config custom.ini

For more information, see README.md
        """,
    )
    parser.add_argument("input", help="Path to input Excel file")
    parser.add_argument("-o", "--output", help="Output Excel report path (default: <input>_report.xlsx)")
    parser.add_argument("--config", help="Path to config.ini (default: ./config.ini)")
    parser.add_argument("--alpha", type=float, help="Weight for answer_correctness (overrides config)")
    parser.add_argument("--beta", type=float, help="Weight for faithfulness (overrides config)")
    parser.add_argument("--gamma", type=float, help="Weight for answer_relevancy (overrides config)")
    parser.add_argument("--model", help="Azure OpenAI deployment name (overrides config)")
    parser.add_argument("--temperature", type=float, help="LLM temperature (overrides config)")
    parser.add_argument("--debug", action="store_true", help="Enable debug logging")
    args = parser.parse_args()

    setup_logging(debug=args.debug)

    if not os.path.exists(args.input):
        logger.error(f"File not found: {args.input}")
        sys.exit(1)
    
    if not args.input.lower().endswith(('.xlsx', '.xls')):
        logger.error(f"Input file must be Excel format (.xlsx or .xls), got: {args.input}")
        sys.exit(1)

    output = args.output or args.input.rsplit(".", 1)[0] + "_report.xlsx"
    
    # Check if output file is writable
    output_dir = os.path.dirname(os.path.abspath(output))
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir, exist_ok=True)
        except OSError as e:
            logger.error(f"Cannot create output directory {output_dir}: {e}")
            sys.exit(1)
    
    # Check if file exists and is writable (could be open in Excel)
    if os.path.exists(output):
        try:
            with open(output, 'a'):
                pass
        except (IOError, OSError):
            logger.error(f"Output file {output} is not writable (may be open in another program)")
            sys.exit(1)

    # Load environment variables - try multiple locations
    # 1. Try .env in same directory as script
    env_local = os.path.join(SCRIPT_DIR, ".env")
    # 2. Try backend/.env if part of larger project
    env_backend = os.path.join(SCRIPT_DIR, "..", "backend", ".env")
    # 3. Try .env in current working directory
    env_cwd = os.path.join(os.getcwd(), ".env")
    
    env_loaded = False
    for env_path in [env_local, env_backend, env_cwd]:
        if os.path.exists(env_path):
            load_dotenv(env_path)
            logger.info(f"Loaded environment from: {env_path}")
            env_loaded = True
            break
    
    if not env_loaded:
        load_dotenv()  # Try default locations
        logger.info("No .env file found, using system environment variables")
    
    # Debug: Check if credentials are loaded
    logger.debug(f"AZURE_OPENAI_ENDPOINT: {os.getenv('AZURE_OPENAI_ENDPOINT', 'NOT SET')}")
    logger.debug(f"AZURE_OPENAI_API_KEY: {'SET' if os.getenv('AZURE_OPENAI_API_KEY') else 'NOT SET'}")

    cfg = load_config(args.config)

    model = args.model or get_cfg(cfg, "azure", "deployment", "gpt-4o")
    temp = args.temperature if args.temperature is not None else get_cfg_float(cfg, "azure", "temperature", 0.0)
    max_rows = get_cfg_int(cfg, "evaluation", "max_rows", 200)
    tox_threshold = get_cfg_float(cfg, "toxicity", "threshold", 0.5)
    ctx_delimiter = get_cfg(cfg, "context", "delimiter", "auto")
    bot_prefix = get_cfg(cfg, "bots", "strip_prefix", "Bot_")
    diag_enabled = get_cfg_bool(cfg, "diagnostics", "enabled", True)
    cache_enabled = get_cfg_bool(cfg, "cache", "enabled", False)

    print("=" * 60)
    print("  RAG Eval — Standalone Report Generator")
    print("=" * 60)
    print(f"  Input:        {args.input}")
    print(f"  Output:       {output}")
    print(f"  Model:        {model}")
    print(f"  Toxicity:     threshold={tox_threshold}")
    print(f"  Context:      delimiter={ctx_delimiter}")
    print(f"  Diagnostics:  {diag_enabled}")
    print(f"  Cache:        {cache_enabled}")
    print(f"  Max rows:     {max_rows}")
    print("-" * 60)

    print("\n[1/3] Parsing input Excel...")
    cases = parse_excel(args.input, max_rows=max_rows,
                        context_delimiter=ctx_delimiter, bot_prefix=bot_prefix)

    print("\n[2/3] Running evaluation...")
    evaluator = StandaloneRagEvaluator(
        cfg=cfg, alpha=args.alpha, beta=args.beta, gamma=args.gamma,
        model_name=model, temperature=temp,
    )

    print(f"  Weights (normalized):")
    print(f"    answer_correctness={evaluator.w_ac:.3f}  faithfulness={evaluator.w_f:.3f}  "
          f"answer_relevancy={evaluator.w_ar:.3f}")
    print(f"    context_precision={evaluator.w_cp:.3f}   context_recall={evaluator.w_cr:.3f}")

    start = time.time()
    results = evaluator.run(cases)
    elapsed = time.time() - start

    print(f"\n[3/3] Writing report...")
    write_report(results, cases, output, toxicity_threshold=tox_threshold,
                 diagnostics_enabled=diag_enabled,
                 metric_thresholds=evaluator.metric_thresholds,
                 llm=evaluator.llm)
    
    # Generate comprehensive JSON report automatically
    print("  Generating comprehensive JSON report...")
    try:
        from pathlib import Path
        output_path = Path(output)
        base_name = output_path.stem
        output_dir = output_path.parent
        
        # Read the Excel file we just created
        excel_file = pd.ExcelFile(output)
        sheet_names = excel_file.sheet_names
        
        # Create comprehensive JSON with all data
        comprehensive_json = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "version": "1.1.0",
                "description": "Comprehensive RAG Evaluation Results",
                "files": {
                    "input": args.input,
                    "report": output_path.name
                },
                "evaluation": {
                    "model": model,
                    "temperature": temp,
                    "time_seconds": elapsed,
                    "embeddings_mode": get_cfg(cfg, "embeddings", "mode", "local"),
                    "toxicity_threshold": tox_threshold,
                    "diagnostics_enabled": diag_enabled,
                    "cache_enabled": cache_enabled
                }
            },
            "input_data": {
                "total_queries": len(cases),
                "test_cases": []
            },
            "evaluation_results": {},
            "configuration": {},
            "summary": {}
        }
        
        # Add input data (original test cases)
        for case in cases:
            comprehensive_json["input_data"]["test_cases"].append({
                "id": case.id,
                "query": case.query,
                "ground_truth": case.ground_truth,
                "bot_responses": case.bot_responses,
                "bot_contexts": case.bot_contexts
            })
        
        # Add evaluation results (all sheets)
        for sheet_name in sheet_names:
            df = pd.read_excel(output, sheet_name=sheet_name)
            df = df.replace({pd.NA: None, pd.NaT: None})
            df = df.where(pd.notnull(df), None)
            records = df.to_dict(orient='records')
            
            sheet_key = sheet_name.lower().replace(' ', '_').replace('-', '_')
            comprehensive_json["evaluation_results"][sheet_key] = {
                "sheet_name": sheet_name,
                "row_count": len(records),
                "columns": list(df.columns),
                "data": records
            }
            
            # Also save individual sheet JSON
            sheet_json_path = output_dir / f"{base_name}_{sheet_key}.json"
            with open(sheet_json_path, 'w', encoding='utf-8') as f:
                json.dump({
                    "sheet_name": sheet_name,
                    "columns": list(df.columns),
                    "data": records
                }, f, indent=2, ensure_ascii=False)
        
        # Add configuration
        config_dict = {}
        for section in cfg.sections():
            config_dict[section] = dict(cfg.items(section))
        comprehensive_json["configuration"]["config_ini"] = config_dict
        
        # Add summary statistics
        if "bot_summary" in comprehensive_json["evaluation_results"]:
            bot_summary = comprehensive_json["evaluation_results"]["bot_summary"]["data"]
            leaderboard = comprehensive_json["evaluation_results"].get("leaderboard", {}).get("data", [])
            
            comprehensive_json["summary"] = {
                "total_bots": len(bot_summary),
                "total_queries": len(cases),
                "winner": results.get("winner"),
                "top_rqs": leaderboard[0]["Avg RQS"] if leaderboard else None,
                "evaluation_time_seconds": elapsed,
                "bots": {
                    bot["Bot"]: {
                        "avg_rqs": bot["Avg RQS"],
                        "answer_correctness": bot["Answer Correctness"],
                        "faithfulness": bot["Faithfulness"],
                        "answer_relevancy": bot["Answer Relevancy"],
                        "context_precision": bot["Context Precision"],
                        "context_recall": bot["Context Recall"],
                        "total_queries": bot["Total Queries"],
                        "retrieval_failures": bot["Retrieval Failures"],
                        "hallucinations": bot["Hallucinations"],
                        "low_quality": bot["Low Quality"],
                        "toxic_queries": bot["Toxic Queries"]
                    }
                    for bot in bot_summary
                }
            }
        
        # Save comprehensive JSON
        comprehensive_json_path = output_dir / f"{base_name}_comprehensive.json"
        with open(comprehensive_json_path, 'w', encoding='utf-8') as f:
            json.dump(comprehensive_json, f, indent=2, ensure_ascii=False)
        
        # Also save the simpler complete JSON (backward compatibility)
        simple_json = {
            "metadata": {
                "source_file": output_path.name,
                "generated_at": datetime.now().isoformat(),
                "sheets": sheet_names,
                "evaluation_time_seconds": elapsed,
                "model": model,
                "temperature": temp
            },
            "data": comprehensive_json["evaluation_results"]
        }
        complete_json_path = output_dir / f"{base_name}_complete.json"
        with open(complete_json_path, 'w', encoding='utf-8') as f:
            json.dump(simple_json, f, indent=2, ensure_ascii=False)
        
        print(f"  ✓ Excel report: {output_path.name}")
        print(f"  ✓ JSON reports: {len(sheet_names)} individual + 1 complete + 1 comprehensive")
        print(f"  ✓ Comprehensive JSON: {comprehensive_json_path.name} ({comprehensive_json_path.stat().st_size / 1024:.1f} KB)")
    except Exception as e:
        logger.warning(f"Failed to generate JSON reports: {e}")
        print(f"  ⚠ JSON generation failed (Excel report still available)")
        import traceback
        logger.debug(traceback.format_exc())

    # Final summary
    print("-" * 60)
    print(f"  Winner:  {results.get('winner', 'N/A')}")
    tox_flagged = sum(1 for s in results.get("toxicity_scores", {}).values() if s >= tox_threshold)
    if tox_flagged:
        print(f"  Toxic:   {tox_flagged}/{len(cases)} queries flagged")
    if not results.get("has_ground_truth"):
        print("  Note:    No ground truth — context_recall & answer_correctness skipped")
    if diag_enabled:
        all_modes = [m.failure_mode for bm in results["bot_metrics"].values() for m in bm.values()]
        ok_count = sum(1 for fm in all_modes if fm == "OK")
        fail_count = len(all_modes) - ok_count
        if fail_count:
            print(f"  Health:  {ok_count} OK, {fail_count} with issues")
    print(f"  Time:    {elapsed:.1f}s")
    print("=" * 60)


if __name__ == "__main__":
    main()
